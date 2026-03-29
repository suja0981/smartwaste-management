"""
routers/reports.py  —  Phase 5: Reports export.

Endpoints:
  GET /reports/export?format=pdf   →  download PDF report
  GET /reports/export?format=xlsx  →  download Excel workbook
  GET /reports/summary             →  JSON summary (no admin required)

Dependencies:
  pip install reportlab openpyxl
"""

import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from database import get_db, BinDB, RouteHistoryDB, TaskDB, CrewDB
from auth_utils import require_admin

logger = logging.getLogger(__name__)
router = APIRouter()


def _now() -> datetime:
    return datetime.now(timezone.utc)


# ─── Data gathering ───────────────────────────────────────────────────────────

def _gather_report_data(db: Session, days: int = 30) -> dict:
    since = _now() - timedelta(days=days)

    bins = db.query(BinDB).order_by(BinDB.zone_id, BinDB.location).all()
    tasks = db.query(TaskDB).filter(TaskDB.created_at >= since).all()
    routes = db.query(RouteHistoryDB).filter(RouteHistoryDB.completion_date >= since).all()
    crews = db.query(CrewDB).all()

    avg_fill = db.query(func.avg(BinDB.fill_level_percent)).scalar() or 0

    return {
        "generated_at": _now(),
        "period_days": days,
        "bins": bins,
        "tasks": tasks,
        "routes": routes,
        "crews": crews,
        "kpis": {
            "total_bins": len(bins),
            "bins_full": sum(1 for b in bins if b.status == "full"),
            "bins_warning": sum(1 for b in bins if b.status == "warning"),
            "bins_offline": sum(1 for b in bins if b.status == "offline"),
            "avg_fill_level": round(avg_fill, 1),
            "routes_completed": len(routes),
            "tasks_completed": sum(1 for t in tasks if t.status == "completed"),
            "total_distance_km": round(sum(r.total_distance_km for r in routes), 2),
            "total_bins_collected": sum(r.bins_collected for r in routes),
        },
    }


# ─── PDF generation ───────────────────────────────────────────────────────────

def _generate_pdf(data: dict) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="reportlab not installed. Run: pip install reportlab",
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=20, spaceAfter=6)
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, spaceAfter=4)
    normal = styles["Normal"]

    kpis = data["kpis"]
    gen = data["generated_at"].strftime("%Y-%m-%d %H:%M UTC")
    period = data["period_days"]

    story = []

    story.append(Paragraph("Smart Waste Management Report", title_style))
    story.append(Paragraph(f"Generated: {gen}  |  Period: last {period} days", normal))
    story.append(HRFlowable(width="100%", spaceAfter=12))

    story.append(Paragraph("Executive Summary", h2_style))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Bins", str(kpis["total_bins"])],
        ["Average Fill Level", f"{kpis['avg_fill_level']}%"],
        ["Bins Currently Full", str(kpis["bins_full"])],
        ["Bins Warning (>=80%)", str(kpis["bins_warning"])],
        ["Bins Offline", str(kpis["bins_offline"])],
        ["Routes Completed", str(kpis["routes_completed"])],
        ["Total Distance Driven", f"{kpis['total_distance_km']} km"],
        ["Bins Collected", str(kpis["total_bins_collected"])],
        ["Tasks Completed", str(kpis["tasks_completed"])],
    ]
    t = Table(kpi_data, colWidths=[9*cm, 7*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    story.append(Paragraph("Bin Status", h2_style))
    bin_rows = [["Bin ID", "Location", "Fill %", "Status", "Zone"]]
    for b in data["bins"]:
        bin_rows.append([
            b.id,
            b.location[:35] + "..." if len(b.location) > 35 else b.location,
            f"{b.fill_level_percent}%",
            b.status.upper(),
            b.zone_id or "—",
        ])
    bt = Table(bin_rows, colWidths=[2.5*cm, 7*cm, 2*cm, 2.5*cm, 3*cm])
    bt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34a853")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(bt)
    story.append(Spacer(1, 16))

    if data["routes"]:
        story.append(Paragraph(f"Collection Routes (last {period} days)", h2_style))
        route_rows = [["Route ID", "Crew", "Bins", "Distance km", "Time min", "Date"]]
        for r in data["routes"][:30]:
            route_rows.append([
                str(r.route_id)[:12],
                str(r.crew_id)[:12],
                str(r.bins_collected),
                f"{r.total_distance_km:.1f}",
                f"{r.total_time_minutes:.0f}",
                r.completion_date.strftime("%Y-%m-%d") if r.completion_date else "—",
            ])
        rt = Table(route_rows, colWidths=[3*cm, 3*cm, 2*cm, 3*cm, 3*cm, 3*cm])
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fbbc04")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(rt)

    doc.build(story)
    return buffer.getvalue()


# ─── Excel generation ─────────────────────────────────────────────────────────

def _generate_xlsx(data: dict) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    wb = openpyxl.Workbook()

    def _header(ws, row, cols):
        fill = PatternFill("solid", fgColor="1a73e8")
        font = Font(bold=True, color="FFFFFF")
        for col, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Smart Waste Management Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A3"] = f"Period: last {data['period_days']} days"

    _header(ws1, 5, ["Metric", "Value"])
    kpis = data["kpis"]
    rows = [
        ("Total Bins", kpis["total_bins"]),
        ("Average Fill Level (%)", kpis["avg_fill_level"]),
        ("Bins Full", kpis["bins_full"]),
        ("Bins Warning", kpis["bins_warning"]),
        ("Bins Offline", kpis["bins_offline"]),
        ("Routes Completed", kpis["routes_completed"]),
        ("Total Distance (km)", kpis["total_distance_km"]),
        ("Total Bins Collected", kpis["total_bins_collected"]),
        ("Tasks Completed", kpis["tasks_completed"]),
    ]
    for i, (k, v) in enumerate(rows, 6):
        ws1.cell(i, 1, k)
        ws1.cell(i, 2, v)
    _auto_width(ws1)

    ws2 = wb.create_sheet("Bins")
    _header(ws2, 1, ["Bin ID", "Location", "Fill %", "Status", "Battery %", "Temp C", "Zone", "Last Telemetry"])
    for i, b in enumerate(data["bins"], 2):
        ws2.cell(i, 1, b.id)
        ws2.cell(i, 2, b.location)
        ws2.cell(i, 3, b.fill_level_percent)
        ws2.cell(i, 4, b.status)
        ws2.cell(i, 5, b.battery_percent)
        ws2.cell(i, 6, b.temperature_c)
        ws2.cell(i, 7, b.zone_id or "")
        ws2.cell(i, 8, b.last_telemetry.strftime("%Y-%m-%d %H:%M") if b.last_telemetry else "")
    _auto_width(ws2)

    ws3 = wb.create_sheet("Route History")
    _header(ws3, 1, ["Route ID", "Crew ID", "Bins Collected", "Distance km", "Time min", "Efficiency", "Date"])
    for i, r in enumerate(data["routes"], 2):
        ws3.cell(i, 1, r.route_id)
        ws3.cell(i, 2, r.crew_id)
        ws3.cell(i, 3, r.bins_collected)
        ws3.cell(i, 4, round(r.total_distance_km, 2))
        ws3.cell(i, 5, round(r.total_time_minutes, 1))
        ws3.cell(i, 6, round(r.fuel_efficiency_score or 0, 3))
        ws3.cell(i, 7, r.completion_date.strftime("%Y-%m-%d") if r.completion_date else "")
    _auto_width(ws3)

    ws4 = wb.create_sheet("Tasks")
    _header(ws4, 1, ["Task ID", "Title", "Priority", "Status", "Location", "Crew", "Created", "Completed"])
    for i, t in enumerate(data["tasks"], 2):
        ws4.cell(i, 1, t.id)
        ws4.cell(i, 2, t.title)
        ws4.cell(i, 3, t.priority)
        ws4.cell(i, 4, t.status)
        ws4.cell(i, 5, t.location)
        ws4.cell(i, 6, t.crew_id or "")
        ws4.cell(i, 7, t.created_at.strftime("%Y-%m-%d") if t.created_at else "")
        ws4.cell(i, 8, t.completed_at.strftime("%Y-%m-%d") if t.completed_at else "")
    _auto_width(ws4)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/export")
def export_report(
    format: str = Query("xlsx", regex="^(pdf|xlsx)$", description="pdf or xlsx"),
    days: int = Query(30, ge=1, le=365, description="Reporting period in days"),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    """
    Export a management report. Admin only.

    GET /reports/export?format=pdf&days=30
    GET /reports/export?format=xlsx&days=7
    """
    data = _gather_report_data(db, days=days)
    ts = data["generated_at"].strftime("%Y%m%d_%H%M")

    if format == "pdf":
        content = _generate_pdf(data)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="waste_report_{ts}.pdf"'},
        )
    else:
        content = _generate_xlsx(data)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="waste_report_{ts}.xlsx"'},
        )


@router.get("/summary")
def get_summary(
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
):
    """
    JSON summary for dashboards or custom rendering.
    No admin required — read-only KPIs.
    GET /reports/summary?days=7
    """
    data = _gather_report_data(db, days=days)
    return {
        "generated_at": data["generated_at"].isoformat(),
        "period_days": days,
        "kpis": data["kpis"],
        "zones": _zone_breakdown(data["bins"]),
    }


def _zone_breakdown(bins) -> dict:
    result = {}
    for b in bins:
        zone = b.zone_id or "unassigned"
        if zone not in result:
            result[zone] = {"total": 0, "full": 0, "fills": []}
        result[zone]["total"] += 1
        result[zone]["fills"].append(b.fill_level_percent)
        if b.status == "full":
            result[zone]["full"] += 1
    for zone, v in result.items():
        fills = v.pop("fills")
        v["avg_fill"] = round(sum(fills) / len(fills), 1) if fills else 0
    return result